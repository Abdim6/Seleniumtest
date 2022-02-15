"l'objectif est de se connecter sur un compte dejà existant"
"Recupère les données des ID via un fichier excel"
"Saisi la date de connexion dans ce fichier"

from asyncio import sleep
from importlib_metadata import email
from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import random
import string
from selenium.webdriver.support.ui import WebDriverWait								
from selenium.webdriver.support import expected_conditions as EC 
import openpyxl
import datetime
import test_driver  

# "Date du jour"
# myDatetime = datetime.datetime.now()
# myString_date = myDatetime.strftime('%d-%m-%Y %H:%M')

# "Recupère les données depuis un fichier excel"
# wb = openpyxl.load_workbook("données_users.xlsx", data_only=True)
# sheet = wb.active
# # email=str(sheet.cell(2,3).value)
# # mdp=str(sheet.cell(2,4).value)
# # sheet.cell(2,6).value = myString_date
# # wb.save("données_users.xlsx")



def seconnecter(id1,id2):
    # "Se rendre sur le site web"
    # driver = webdriver.Chrome()
    # time.sleep(1)
    # driver.maximize_window()
    # driver.get("http://demostore.supersqa.com/")
    # time.sleep(2)
    driver = test_driver.driver
    myaccount_btn = driver.find_element(By.XPATH, '//*[@id="site-navigation"]/div[1]/ul/li[4]/a')
    myaccount_btn.click()
    time.sleep(2)
    "Début de la connexion"
    username_btn = driver.find_element(By.ID, "username")
    username_btn.send_keys(id1)
    time.sleep(2)
    pwd_btn = driver.find_element(By.ID, "password")
    pwd_btn.send_keys(id2)
    time.sleep(2)
    validate_btn = driver.find_element(By.XPATH,'//*[@id="customer_login"]/div[1]/form/p[3]/button').click()
    time.sleep(2)

    "Vérification finale de la connexion de compte, ceci vérifie si le user a atteind la page voulue"
    wait = WebDriverWait(driver,10)
    logout_btn = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, '#post-9 > div > div > nav > ul > li.woocommerce-MyAccount-navigation-link.woocommerce-MyAccount-navigation-link--customer-logout > a')))	
    time.sleep(2)
    print("La connexion est OK")

    # driver.quit()
    
# for i in range (sheet.max_row-1):
#     email=str(sheet.cell(2+i,3).value)
#     mdp=str(sheet.cell(2+i,4).value)
#     seconnecter(email,mdp)
#     "Date du jour"
#     myDatetime = datetime.datetime.now()
#     myString_date = myDatetime.strftime('%d-%m-%Y %H:%M')
#     sheet.cell(2+i,6).value = myString_date
#     wb.save("données_users.xlsx")
#     print(f"Connexion user {i+1} - {email} est Ok")