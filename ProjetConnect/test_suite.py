"c'est un fichier qui ping le fichier excel pour récupérer les données"
"Ajouter le resultat du test dans le fichier + la date d'exécution"
""

from asyncio import sleep
from importlib_metadata import email
from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import random
import string
from selenium.webdriver.support.ui import WebDriverWait								
from selenium.webdriver.support import expected_conditions as EC 
import openpyxl
import datetime
# from test_connexion2 import seconnecter
import test_connexion2
from test_deco import sedeconnecter

"Date du jour"
# myDatetime = datetime.datetime.now()
# myString_date = myDatetime.strftime('%d-%m-%Y %H:%M')

"Recupère les données depuis un fichier excel"
wb = openpyxl.load_workbook("données_users2.xlsx", data_only=True)
sheet = wb.active
# email=str(sheet.cell(2,3).value)
# mdp=str(sheet.cell(2,4).value)
# sheet.cell(2,6).value = myString_date
# wb.save("données_users.xlsx")

for i in range (sheet.max_row-1):
    email=str(sheet.cell(2+i,3).value)
    mdp=str(sheet.cell(2+i,4).value)
    print("------------------")
    test_connexion2.seconnecter(email,mdp)
    "Date du jour"
    myDatetime = datetime.datetime.now()
    myString_date = myDatetime.strftime('%d-%m-%Y %H:%M')
    sheet.cell(2+i,6).value = myString_date
    wb.save("données_users2.xlsx")
    print(f"Connexion user {i+1} - {email} est Ok")
    sedeconnecter()
    time.sleep(2)
    
# email=str(sheet.cell(2,3).value)
# mdp=str(sheet.cell(2,4).value)
# seconnecter(email,mdp)
# "Date du jour"
# myDatetime = datetime.datetime.now()
# myString_date = myDatetime.strftime('%d-%m-%Y %H:%M')
# sheet.cell(2,6).value = myString_date
# wb.save("données_users2.xlsx")
# print(f"Connexion user {+1} - {email} est Ok")
# sedeconnecter()
# time.sleep(2)