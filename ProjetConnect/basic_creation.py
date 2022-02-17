"Faire la souscription avec des ID random"
"La connexion à la chaine en boucle for avec tous les users"
"Fais au moins une action "
"La deconnexion doit se faire après l'exé de la connexion"
"Pense à des temps de latence"

from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import random
import string
from selenium.webdriver.support.ui import WebDriverWait								
from selenium.webdriver.support import expected_conditions as EC 
import openpyxl
import datetime
import logging 

logging.basicConfig(filename="//Users//abdi.bileh17//Documents//Selenium//SeleniumTuto//ProjetConnect//testLog.log") 

logging.debug("---- DEBUT DE TEST----")
# driver = webdriver.Chrome()
# driver.get("http://demostore.supersqa.com/my-account/")

"Generation de l'email"
letters = string.ascii_lowercase
rand_string = "".join(random.choice(letters) for i in range(15))
rand_mail = rand_string + "@gmail.com"
print("L'email a rentrer est : "+rand_mail)

"Generation de mdp"
mdp=string.ascii_letters
# letters = string.ascii_lowercase
rand_mdp = "".join(random.choice(mdp) for i in range(15))
print("Le mot de passe choisi est : "+rand_mdp)

"prenom & nom"
prenom = rand_string[1:7]
mdp = rand_string[7:]

"Date du jour"
myDatetime = datetime.datetime.now()
myString_date = myDatetime.strftime('%d-%m-%Y %H:%M')
"-------------"
# Saisir dans le fichier les ID et les mdp du nouveau user
"-------------"


"Ecrire les données dans le fichier excel"
wb = openpyxl.load_workbook("donneesUsers.xlsx", data_only=True)
sheet = wb.active
x = sheet.max_row +1
sheet.cell(x,1).value = prenom
sheet.cell(x,2).value = mdp
sheet.cell(x,3).value = rand_mail
sheet.cell(x,4).value = rand_mdp
sheet.cell(x,5).value = myString_date

wb.save("donneesUsers.xlsx")


"Se rendre sur le site web"
driver = webdriver.Chrome()
time.sleep(1)
driver.maximize_window()
driver.get("http://demostore.supersqa.com/")
time.sleep(2)

logging.info("Accès au site - OK")
myaccount_btn = driver.find_element(By.XPATH, '//*[@id="site-navigation"]/div[1]/ul/li[4]/a')
myaccount_btn.click()
time.sleep(2)

"Début de la création de compte"
id_section = driver.find_element(By.ID,"reg_email")
pwd_section = driver.find_element(By.ID, "reg_password")
id_section.clear()
# id_section.send_keys("my_email")
id_section.send_keys(rand_mail)
pwd_section.send_keys(rand_mdp)
time.sleep(3)


logging.info("Saisi du mail & de mot de passe - OK")

Submit_btn = driver.find_element(By.XPATH, '//*[@id="customer_login"]/div[2]/form/p[3]/button')
Submit_btn.click()

"Vérification finale de la création de compte, ceci vérifie si le user a atteind la page voulue"
wait = WebDriverWait(driver,10)
logout_btn = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, '#post-9 > div > div > nav > ul > li.woocommerce-MyAccount-navigation-link.woocommerce-MyAccount-navigation-link--customer-logout > a')))	
time.sleep(2)

logging.info("Création de compte est - OK")
print("PASS")

logging.info("---- FIN DE TEST----")
			







# "register"
# # username_btn = driver.find_element(By.ID, "reg_email")
# # username_btn.send_keys("abdemdsd")
# # time.sleep(2)
# # pwd_btn = driver.find_element(By.ID, "reg_password")
# # pwd_btn.send_keys("azertyuiop")
# # time.sleep(2)
# # validate_btn = driver.find_element(By.XPATH,'//*[@id="customer_login"]/div[2]/form/p[3]/button').click()
# # time.sleep(2)


driver.quit()

